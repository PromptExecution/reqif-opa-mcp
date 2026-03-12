"""Deterministic XLSX extraction and distillation."""

from __future__ import annotations

import re
from pathlib import Path
from typing import TYPE_CHECKING, Any

from openpyxl import load_workbook  # type: ignore[import-untyped]
from returns.result import Failure, Result, Success

from reqif_ingest_cli.artifact import register_artifact
from reqif_ingest_cli.models import DocumentGraph, DocumentNode, RequirementCandidate, SourceAnchor
from reqif_ingest_cli.utils import (
    column_name,
    map_header_name,
    normalize_text,
    split_paragraphs,
    stable_id,
)

if TYPE_CHECKING:
    from openpyxl.workbook.workbook import Workbook  # type: ignore[import-untyped]
    from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import-untyped]
else:
    Workbook = Any
    Worksheet = Any

_AESC_SHEET_PATTERN = re.compile(r"^[A-Z-]+(?:\d+|AP)$")
_PRACTICE_ID_PATTERN = re.compile(r"^[A-Z-]+-\d+[a-z]?$")


def extract_xlsx_document(
    path: str | Path,
    source_uri: str | None = None,
    profile: str = "auto",
) -> Result[DocumentGraph, Exception]:
    """Extract a workbook into the canonical document graph."""
    workbook: Workbook | None = None
    try:
        resolved = Path(path).expanduser().resolve()
        workbook = load_workbook(resolved, read_only=True, data_only=True)
        resolved_profile = detect_xlsx_profile(workbook) if profile == "auto" else profile

        artifact_result = register_artifact(
            resolved,
            source_uri=source_uri,
            document_profile=resolved_profile,
        )
        if isinstance(artifact_result, Failure):
            return artifact_result
        artifact = artifact_result.unwrap()

        if resolved_profile == "aescsf_core_v2":
            graph = _extract_aescsf_core(workbook, artifact)
        elif resolved_profile == "aescsf_toolkit_v1_1":
            graph = _extract_aescsf_toolkit(workbook, artifact)
        else:
            graph = _extract_generic_workbook(workbook, artifact)
        return Success(graph)
    except Exception as exc:
        return Failure(exc)
    finally:
        if workbook is not None:
            workbook.close()


def distill_xlsx_requirements(
    path: str | Path,
    source_uri: str | None = None,
    profile: str = "auto",
) -> Result[list[RequirementCandidate], Exception]:
    """Distill deterministic requirement candidates from a workbook."""
    graph_result = extract_xlsx_document(path, source_uri=source_uri, profile=profile)
    if isinstance(graph_result, Failure):
        return graph_result
    return Success(distill_xlsx_graph(graph_result.unwrap()))


def detect_xlsx_profile(workbook: Workbook) -> str:
    """Detect the workbook profile from stable sheet structure."""
    first_sheet = workbook[workbook.sheetnames[0]]
    try:
        header_row = next(first_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    except StopIteration:
        # Empty sheet: no header row available, fall back to generic profile.
        return "generic_xlsx_table"
    first_row = [normalize_text(value) for value in header_row]
    header_set = set(first_row)
    if {
        "Domain",
        "Objective ID",
        "Objective",
        "Practice ID",
        "Practice",
    }.issubset(header_set):
        return "aescsf_core_v2"

    practice_sheet_count = 0
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        if _AESC_SHEET_PATTERN.fullmatch(sheet_name) and normalize_text(worksheet.cell(4, 4).value) == "Practices":
            practice_sheet_count += 1
    if practice_sheet_count:
        return "aescsf_toolkit_v1_1"

    return "generic_xlsx_table"


def distill_xlsx_graph(graph: DocumentGraph) -> list[RequirementCandidate]:
    """Create deterministic requirement candidates from workbook row records."""
    candidates: list[RequirementCandidate] = []
    child_nodes: dict[str, list[DocumentNode]] = {}
    for node in graph.nodes:
        if node.parent_id is not None:
            child_nodes.setdefault(node.parent_id, []).append(node)

    for node in graph.nodes:
        if node.node_type != "row_record":
            continue
        text = str(node.attributes.get("practice") or node.attributes.get("requirement_text") or node.text or "").strip()
        if not text:
            continue

        key = str(
            node.attributes.get("practice_id")
            or node.attributes.get("requirement_id")
            or node.semantic_id
            or node.node_id
        )
        section = _optional_text(node.attributes.get("objective_id") or node.attributes.get("worksheet"))
        if graph.profile == "aescsf_core_v2":
            extraction_rule_id = "xlsx.aescsf_core.practice_row.v1"
            rationale = "Mapped Practice ID and Practice columns from the AESCSF core workbook."
        elif graph.profile == "aescsf_toolkit_v1_1":
            extraction_rule_id = "xlsx.aescsf_toolkit.practice_sheet.v1"
            rationale = "Mapped practice rows from the AESCSF assessment toolkit workbook layout."
        else:
            extraction_rule_id = "xlsx.generic.requirement_row.v1"
            rationale = "Mapped requirement-like rows from a detected workbook header layout."

        source_children = child_nodes.get(node.node_id, [])
        source_node_ids = [node.node_id, *[child.node_id for child in source_children]]
        anchors = list(node.anchors)
        for child in source_children:
            anchors.extend(child.anchors)

        metadata = dict(node.attributes)
        context_guidance = [
            child.text
            for child in source_children
            if child.attributes.get("role") == "context_guidance" and child.text
        ]
        if context_guidance:
            metadata["context_guidance"] = context_guidance

        subtype_hints = [
            value
            for value in [
                _optional_text(node.attributes.get("domain")),
                _optional_text(node.attributes.get("security_profile")),
            ]
            if value
        ]

        candidates.append(
            RequirementCandidate(
                schema="requirement_candidate/1",
                candidate_id=stable_id("candidate", graph.artifact.artifact_id, key, text[:96]),
                artifact_id=graph.artifact.artifact_id,
                artifact_sha256=graph.artifact.sha256,
                profile=graph.profile,
                key=key,
                text=text,
                section=section,
                subtype_hints=subtype_hints,
                extraction_rule_id=extraction_rule_id,
                rationale=rationale,
                confidence_source="deterministic",
                source_node_ids=source_node_ids,
                anchors=anchors,
                metadata=metadata,
            )
        )

    return candidates


def _extract_aescsf_core(workbook: Workbook, artifact: Any) -> DocumentGraph:
    """Extract the flat AESCSF core workbook."""
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(min_row=1, max_row=1, values_only=True)
    headers = [normalize_text(value) for value in next(rows)]
    semantic_headers = {
        index + 1: map_header_name(header, index + 1)
        for index, header in enumerate(headers)
    }
    header_map = {
        column_name(index): semantic_headers[index]
        for index in semantic_headers
    }
    nodes: list[DocumentNode] = [
        _worksheet_node(artifact.artifact_id, worksheet.title, worksheet.title)
    ]

    for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        row_values = {
            semantic_headers[index + 1]: normalize_text(value)
            for index, value in enumerate(row)
            if semantic_headers.get(index + 1)
        }
        practice_id = row_values.get("practice_id", "")
        practice = row_values.get("practice", "")
        if not practice_id or not practice:
            continue

        row_id = stable_id("row", artifact.artifact_id, worksheet.title, row_index, practice_id)
        nodes.append(
            DocumentNode(
                node_id=row_id,
                node_type="row_record",
                text=practice,
                parent_id=stable_id("worksheet", artifact.artifact_id, worksheet.title),
                semantic_id=practice_id,
                attributes={
                    **{key: value for key, value in row_values.items() if value},
                    "worksheet": worksheet.title,
                },
                anchors=[
                    SourceAnchor(
                        kind="xlsx_row",
                        artifact_id=artifact.artifact_id,
                        sheet=worksheet.title,
                        row=row_index,
                        semantic_id=practice_id,
                        heading_path=[row_values.get("objective_id", "")] if row_values.get("objective_id") else [],
                    )
                ],
            )
        )
        nodes.extend(
            _build_row_paragraph_nodes(
                artifact_id=artifact.artifact_id,
                worksheet=worksheet.title,
                row_index=row_index,
                row_values=row_values,
                row_id=row_id,
            )
        )

    return DocumentGraph(
        schema="document_graph/1",
        artifact=artifact,
        profile="aescsf_core_v2",
        nodes=nodes,
        metadata={"header_map": {worksheet.title: header_map}},
    )


def _extract_aescsf_toolkit(workbook: Workbook, artifact: Any) -> DocumentGraph:
    """Extract the multi-sheet AESCSF assessment toolkit."""
    nodes: list[DocumentNode] = []
    metadata: dict[str, Any] = {"sheets": []}

    for worksheet in workbook.worksheets:
        if not _is_aescsf_toolkit_sheet(worksheet):
            continue

        objective_id = normalize_text(worksheet.cell(2, 4).value)
        objective = normalize_text(worksheet.cell(3, 4).value)
        worksheet_node = _worksheet_node(artifact.artifact_id, worksheet.title, objective or worksheet.title)
        nodes.append(worksheet_node)
        metadata["sheets"].append(
            {
                "sheet": worksheet.title,
                "objective_id": objective_id,
                "objective": objective,
            }
        )

        row_index = 1
        while row_index <= worksheet.max_row:
            practice_id = normalize_text(worksheet.cell(row_index, 1).value)
            practice = normalize_text(worksheet.cell(row_index, 4).value)
            if not _PRACTICE_ID_PATTERN.fullmatch(practice_id) or not practice:
                row_index += 1
                continue

            mil = normalize_text(worksheet.cell(row_index, 10).value)
            security_profile = normalize_text(worksheet.cell(row_index, 12).value)
            row_id = stable_id("row", artifact.artifact_id, worksheet.title, row_index, practice_id)
            row_values = {
                "objective_id": objective_id,
                "objective": objective,
                "practice_id": practice_id,
                "practice": practice,
                "mil": mil,
                "security_profile": security_profile,
                "worksheet": worksheet.title,
            }
            nodes.append(
                DocumentNode(
                    node_id=row_id,
                    node_type="row_record",
                    text=practice,
                    parent_id=worksheet_node.node_id,
                    semantic_id=practice_id,
                    attributes={key: value for key, value in row_values.items() if value},
                    anchors=[
                        SourceAnchor(
                            kind="xlsx_row",
                            artifact_id=artifact.artifact_id,
                            sheet=worksheet.title,
                            row=row_index,
                            semantic_id=practice_id,
                            heading_path=[objective_id] if objective_id else [],
                        )
                    ],
                )
            )
            nodes.extend(
                _build_paragraph_nodes_for_cell(
                    artifact_id=artifact.artifact_id,
                    worksheet=worksheet.title,
                    row_index=row_index,
                    column_index=4,
                    role="practice",
                    semantic_id=practice_id,
                    text=practice,
                    parent_id=row_id,
                    heading_path=[objective_id] if objective_id else [],
                )
            )

            cursor = row_index + 1
            saw_context_label = False
            while cursor <= worksheet.max_row:
                next_practice_id = normalize_text(worksheet.cell(cursor, 1).value)
                if _PRACTICE_ID_PATTERN.fullmatch(next_practice_id):
                    break
                candidate_text = normalize_text(worksheet.cell(cursor, 4).value)
                if candidate_text == "Context and Guidance":
                    saw_context_label = True
                elif saw_context_label and candidate_text:
                    nodes.extend(
                        _build_paragraph_nodes_for_cell(
                            artifact_id=artifact.artifact_id,
                            worksheet=worksheet.title,
                            row_index=cursor,
                            column_index=4,
                            role="context_guidance",
                            semantic_id=practice_id,
                            text=candidate_text,
                            parent_id=row_id,
                            heading_path=[objective_id] if objective_id else [],
                        )
                    )
                cursor += 1
            row_index = cursor

    return DocumentGraph(
        schema="document_graph/1",
        artifact=artifact,
        profile="aescsf_toolkit_v1_1",
        nodes=nodes,
        metadata=metadata,
    )


def _extract_generic_workbook(workbook: Workbook, artifact: Any) -> DocumentGraph:
    """Extract a generic header-based workbook."""
    nodes: list[DocumentNode] = []
    metadata: dict[str, Any] = {"header_map": {}}

    for worksheet in workbook.worksheets:
        header_row_index = _detect_header_row(worksheet)
        headers = [
            normalize_text(value)
            for value in next(
                worksheet.iter_rows(
                    min_row=header_row_index,
                    max_row=header_row_index,
                    values_only=True,
                )
            )
        ]
        semantic_headers = {
            index + 1: map_header_name(header, index + 1)
            for index, header in enumerate(headers)
        }
        metadata["header_map"][worksheet.title] = {
            column_name(index): semantic_headers[index]
            for index in semantic_headers
        }
        worksheet_node = _worksheet_node(artifact.artifact_id, worksheet.title, worksheet.title)
        nodes.append(worksheet_node)

        for row_index, row in enumerate(
            worksheet.iter_rows(min_row=header_row_index + 1, values_only=True),
            start=header_row_index + 1,
        ):
            row_values = {
                semantic_headers[index + 1]: normalize_text(value)
                for index, value in enumerate(row)
                if semantic_headers.get(index + 1)
            }
            if not any(row_values.values()):
                continue

            text = row_values.get("requirement_text") or row_values.get("practice")
            key = row_values.get("requirement_id") or row_values.get("practice_id")
            if not text:
                continue

            row_id = stable_id("row", artifact.artifact_id, worksheet.title, row_index, key or text[:32])
            nodes.append(
                DocumentNode(
                    node_id=row_id,
                    node_type="row_record",
                    text=text,
                    parent_id=worksheet_node.node_id,
                    semantic_id=key or stable_id("semantic", worksheet.title, row_index, text[:64]),
                    attributes={
                        **{k: v for k, v in row_values.items() if v},
                        "worksheet": worksheet.title,
                    },
                    anchors=[
                        SourceAnchor(
                            kind="xlsx_row",
                            artifact_id=artifact.artifact_id,
                            sheet=worksheet.title,
                            row=row_index,
                            semantic_id=key or None,
                        )
                    ],
                )
            )
            nodes.extend(
                _build_row_paragraph_nodes(
                    artifact_id=artifact.artifact_id,
                    worksheet=worksheet.title,
                    row_index=row_index,
                    row_values=row_values,
                    row_id=row_id,
                )
            )

    return DocumentGraph(
        schema="document_graph/1",
        artifact=artifact,
        profile="generic_xlsx_table",
        nodes=nodes,
        metadata=metadata,
    )


def _worksheet_node(artifact_id: str, sheet_name: str, label: str) -> DocumentNode:
    """Create a worksheet node."""
    return DocumentNode(
        node_id=stable_id("worksheet", artifact_id, sheet_name),
        node_type="worksheet",
        text=label,
        parent_id=None,
        semantic_id=sheet_name,
        attributes={"worksheet": sheet_name},
        anchors=[
            SourceAnchor(
                kind="xlsx_sheet",
                artifact_id=artifact_id,
                sheet=sheet_name,
                semantic_id=sheet_name,
            )
        ],
    )


def _build_row_paragraph_nodes(
    artifact_id: str,
    worksheet: str,
    row_index: int,
    row_values: dict[str, Any],
    row_id: str,
) -> list[DocumentNode]:
    """Create paragraph nodes for every populated cell in a row.

    The ``row_values`` mapping is keyed by semantic header names. Values may be either:

    * a single string (legacy behavior), or
    * a list of ``(column_index, text)`` tuples, allowing multiple columns to map
      to the same semantic header while preserving the original column indices.
    """
    nodes: list[DocumentNode] = []

    # These identifiers are per-row and shared across all paragraph nodes.
    semantic_id = row_values.get("practice_id") or row_values.get("requirement_id")
    objective_id = row_values.get("objective_id")
    heading_path = [objective_id] if objective_id else []

    # Fallback position counter used only when no explicit column index is provided.
    position = 1

    for semantic_header, value in row_values.items():
        # Skip non-content keys that are used solely for identification or headings.
        if semantic_header in {"practice_id", "requirement_id", "objective_id"}:
            continue

        # Normalize the value to a list of (column_index, text) pairs.
        # This supports both legacy string values and the richer representation
        # where multiple columns map to the same semantic header.
        if isinstance(value, list):
            cells = value
        else:
            cells = [(None, value)]

        for explicit_column_index, text in cells:
            if not text:
                continue

            role = "context_guidance" if semantic_header == "context_guidance" else semantic_header

            # Prefer an explicit column index if provided; otherwise, fall back
            # to a monotonically increasing position counter to preserve the
            # previous behavior when no column information is available.
            column_index = explicit_column_index if explicit_column_index is not None else position

            nodes.extend(
                _build_paragraph_nodes_for_cell(
                    artifact_id=artifact_id,
                    worksheet=worksheet,
                    row_index=row_index,
                    column_index=column_index,
                    role=role,
                    semantic_id=semantic_id,
                    text=text,
                    parent_id=row_id,
                    heading_path=heading_path,
                )
            )

            # Only advance the fallback position counter when it was actually used.
            if explicit_column_index is None:
                position += 1
    return nodes


def _build_paragraph_nodes_for_cell(
    artifact_id: str,
    worksheet: str,
    row_index: int,
    column_index: int,
    role: str,
    semantic_id: str | None,
    text: str,
    parent_id: str,
    heading_path: list[str],
) -> list[DocumentNode]:
    """Chunk a cell into paragraph nodes."""
    cell_ref = f"{column_name(column_index)}{row_index}"
    nodes: list[DocumentNode] = []
    for paragraph_index, paragraph in enumerate(split_paragraphs(text), start=1):
        node_id = stable_id(
            "paragraph",
            artifact_id,
            worksheet,
            cell_ref,
            paragraph_index,
            paragraph[:96],
        )
        nodes.append(
            DocumentNode(
                node_id=node_id,
                node_type="paragraph",
                text=paragraph,
                parent_id=parent_id,
                semantic_id=semantic_id,
                attributes={"role": role, "cell": cell_ref},
                anchors=[
                    SourceAnchor(
                        kind="xlsx_cell_paragraph",
                        artifact_id=artifact_id,
                        sheet=worksheet,
                        row=row_index,
                        column=column_name(column_index),
                        cell=cell_ref,
                        paragraph=paragraph_index,
                        heading_path=list(heading_path),
                        semantic_id=semantic_id,
                    )
                ],
            )
        )
    return nodes


def _is_aescsf_toolkit_sheet(worksheet: Worksheet) -> bool:
    """Check whether a worksheet matches the assessment toolkit layout."""
    return bool(
        _AESC_SHEET_PATTERN.fullmatch(worksheet.title)
        and normalize_text(worksheet.cell(4, 4).value) == "Practices"
    )


def _detect_header_row(worksheet: Worksheet, search_rows: int = 10) -> int:
    """Pick the most header-like row near the top of the sheet."""
    best_row = 1
    best_score = -1
    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=search_rows, values_only=True),
        start=1,
    ):
        values = [normalize_text(value) for value in row]
        non_empty = [value for value in values if value]
        if not non_empty:
            continue
        # Count cells whose header text maps to a recognized semantic name,
        # i.e. where map_header_name does not return the generic fallback
        # "column_{col_index}" for that column.
        semantic_hits = sum(
            1
            for col_index, value in enumerate(values, start=1)
            if value and map_header_name(value, col_index) != f"column_{col_index}"
        )
        score = len(non_empty) + semantic_hits
        if score > best_score:
            best_row = row_index
            best_score = score
    return best_row


def _optional_text(value: object) -> str | None:
    """Convert a value to optional trimmed text."""
    text = str(value).strip() if value is not None else ""
    return text or None
